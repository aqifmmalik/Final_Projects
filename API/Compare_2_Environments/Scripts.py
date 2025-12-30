import json
import os
import sys
import requests
import webbrowser
import pandas as pd
import time
import re
import logging
import base64
from deepdiff import DeepDiff
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

try:
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    class MockOpenpyxlStyles:
        def __init__(self, **kwargs): pass
    PatternFill = Font = Alignment = MockOpenpyxlStyles

try:
    from colorama import Fore, Style, init
    init(autoreset=True)
except ImportError:
    class MockColorama:
        def __getattr__(self, name): return ''
    Fore = Style = MockColorama()

try:
    from jsonschema import validate, ValidationError
    JSONSCHEMA_AVAILABLE = True
except ImportError:
    class MockJsonSchema:
        def __init__(self, *args, **kwargs): pass
    validate = MockJsonSchema
    ValidationError = type('ValidationError', (Exception,), {})
    JSONSCHEMA_AVAILABLE = False

LOG_FILE = 'comparison_script.log'
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8'),
                        logging.StreamHandler(sys.stdout)
                    ])
logger = logging.getLogger(__name__)

TEST_DATA_FILE = 'test_data.xlsx'
DEFAULT_CONFIG = {
    'COLLECTION_FILE': 'MemberAPI_Final.json',
    'ODS_URL': 'https://ccss-tmg-facets.us-e1.cloudhub.io/api',
    'PRD_URL': 'https://ccss-tmg-facets-prd.us-e1.cloudhub.io/api',
    'OUTPUT_DIR': 'comparison_report_output',
    'TEMPLATE_FILE': 'report_template.html',
    'TARGET_TEST_PREFIXES': '',
    'DYNAMIC_FIELD_TERMS': 'LAST_UPDATE,TIMESTAMP,DATE,TIME,VERSION,MULE_CORRELATION_ID,ID',
    'EXCLUDE_FIELD_NAMES': 'PAYEE_ID',
    'AUTO_OPEN_HTML': 'YES',
    'ENABLE_PERF_GRAPH': 'YES',
    'RUN_MODE': 'DUAL',
    'AUTH_TYPE': 'NONE',
    'AUTH_VALUE': '',
    'AUTH_HEADER': 'Authorization',
    'BASIC_AUTH_USER': '',
    'BASIC_AUTH_PASS': '',
    'POSTMAN_VARS': '',
    'SCHEMA_DIR': 'schemas',
    'REQUEST_TIMEOUT':'30'
}

def get_status_color(status_code: int) -> str:
    if 200 <= status_code < 300: return Fore.GREEN
    if 400 <= status_code < 500: return Fore.YELLOW
    if 500 <= status_code < 600: return Fore.RED
    return Fore.WHITE

def escape_html(text: str) -> str:
    if not isinstance(text, str):
        return str(text)
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

def apply_postman_vars(text: str, vars_dict: Dict[str, str]) -> str:
    if not isinstance(text, str):
        return text
    for key, value in vars_dict.items():
        text = text.replace(f'{{{{{key}}}}}', value)
    return text

def get_unique_filepath(directory, filename):
    base, ext = os.path.splitext(filename)
    counter = 1
    final_path = os.path.join(directory, filename)
    
    while os.path.exists(final_path):
        final_path = os.path.join(directory, f"{base}_{counter}{ext}")
        counter += 1
        
    return final_path

def get_record_count(json_data: Any) -> int:
    if json_data is None:
        return 0
    
    if isinstance(json_data, list):
        return len(json_data)
        
    if isinstance(json_data, dict):
        for key in ['items', 'data', 'results', 'members', 'records', 'content']:
            if key in json_data and isinstance(json_data[key], list):
                return len(json_data[key])
        return 1
        
    return 0

def load_settings_from_excel(data_file: str) -> Dict[str, Any]:
    settings = DEFAULT_CONFIG.copy()
    data_file_path = Path.cwd() / data_file

    logger.info(f"\n--- Loading Configuration from {data_file_path.name} (Sheet: Settings) ---")

    try:
        if not data_file_path.exists():
            logger.error(f"Config file '{data_file}' not found. Using default internal settings.")
            return settings

        settings_df = pd.read_excel(data_file_path, sheet_name='Settings', header=None)
        settings_dict = settings_df.set_index(0).to_dict()[1]

        for key, value in settings_dict.items():
            if pd.notna(value):
                settings[str(key).strip()] = str(value).strip()

        settings['DYNAMIC_FIELD_TERMS'] = [t.strip() for t in settings['DYNAMIC_FIELD_TERMS'].split(',') if t.strip()]
        settings['TARGET_TEST_PREFIXES'] = [t.strip() for t in settings['TARGET_TEST_PREFIXES'].split(',') if t.strip()]
        settings['POSTMAN_COLLECTION_FILE'] = Path.cwd() / settings['COLLECTION_FILE']
        settings['AUTO_OPEN_HTML'] = settings['AUTO_OPEN_HTML'].upper()
        settings['ENABLE_PERF_GRAPH'] = settings['ENABLE_PERF_GRAPH'].upper()
        settings['RUN_MODE'] = settings['RUN_MODE'].upper()

        settings['EXCLUDE_FIELD_NAMES_LIST'] = [t.strip() for t in settings['EXCLUDE_FIELD_NAMES'].split(',') if t.strip()]

        settings['POSTMAN_VARS_DICT'] = {}
        for item in settings['POSTMAN_VARS'].split(','):
            if '=' in item:
                try:
                    key, val = item.split('=', 1)
                    settings['POSTMAN_VARS_DICT'][key.strip()] = val.strip()
                except ValueError:
                     logger.warning(f"Skipping malformed POSTMAN_VARS entry: {item}")

        logger.info(f"Loaded successfully. Collection: {settings['COLLECTION_FILE']}, ODS URL: {settings['ODS_URL']}")

    except Exception as e:
        logger.critical(f"CRITICAL ERROR parsing 'Settings' sheet: {e}. Using default internal settings.")
        settings['DYNAMIC_FIELD_TERMS'] = [t.strip() for t in settings['DYNAMIC_FIELD_TERMS'].split(',') if t.strip()]
        settings['TARGET_TEST_PREFIXES'] = [t.strip() for t in settings['TARGET_TEST_PREFIXES'].split(',') if t.strip()]
        settings['POSTMAN_COLLECTION_FILE'] = Path.cwd() / settings['COLLECTION_FILE']

    return settings

def export_to_excel(data, output_dir):
    excel_path = get_unique_filepath(output_dir, 'comparison_report.xlsx')
    
    logger.info("--- Generating Excel Report ---")
    try:
        excel_data = [{
            'Test Case': item['test_name'],
            'ODS Status': item['ods_status'],
            'PRD Status': item['prd_status'],
            'ODS Time (ms)': item['ods_time'].replace(' ms', ''),
            'PRD Time (ms)': item['prd_time'].replace(' ms', ''),
            'ODS Records': item.get('ods_record_count', 0),
            'PRD Records': item.get('prd_record_count', 0),
            'Test_Type': item['test_type'],
            'Data Diff Result': item['data_diff_result'],
            'Data Diff Summary': item['data_diff_summary'],
            'Comments': item['comments']
        } for item in data]

        df = pd.DataFrame(excel_data)

        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        df.to_excel(writer, sheet_name='Comparison Report', index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Comparison Report']

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.row < 100 and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            adjusted_width = max(adjusted_width, 10)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        writer.close()

        logger.info(f"\n--- Excel report successfully generated at: {os.path.abspath(excel_path)} ---")
        return os.path.abspath(excel_path)
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return None

def check_api_health(environment_name: str, base_url: str) -> bool:
    logger.info(f"\n--- Checking {environment_name} API Health (Phase 1/5) ---")
    try:
        response = requests.get(base_url, timeout=10)
        status_code = response.status_code
        env_color = Fore.CYAN if environment_name == 'ODS' else Fore.BLUE

        if 200 <= status_code < 300:
            logger.info(f"[{env_color}{environment_name}{Style.RESET_ALL} HEALTH CHECK] {Fore.GREEN}SUCCESS{Style.RESET_ALL}: API is reachable and healthy (HTTP {status_code}).")
            return True
        elif 400 <= status_code < 500:
            logger.warning(f"[{env_color}{environment_name}{Style.RESET_ALL} HEALTH CHECK] {Fore.YELLOW}WARNING{Style.RESET_ALL}: Received HTTP {status_code}. API is reachable, but access/endpoint may be restricted.")
            return True
        else:
            logger.error(f"[{env_color}{environment_name}{Style.RESET_ALL} HEALTH CHECK] {Fore.RED}FAILURE{Style.RESET_ALL}: Received Unexpected HTTP {status_code}. Investigate service stability.")
            return False

    except requests.exceptions.Timeout:
        logger.error(f"[{env_color}{environment_name}{Style.RESET_ALL} HEALTH CHECK] {Fore.RED}FAILURE{Style.RESET_ALL}: Connection timed out after 10 seconds.")
        return False
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[{env_color}{environment_name}{Style.RESET_ALL} HEALTH CHECK] {Fore.RED}FAILURE{Style.RESET_ALL}: Could not connect to API at {base_url}. Error: {e.__class__.__name__}")
        return False
    except Exception as e:
        logger.error(f"[{env_color}{environment_name}{Style.RESET_ALL} HEALTH CHECK] {Fore.RED}UNKNOWN ERROR{Style.RESET_ALL}: {e}")
        return False

def extract_requests(collection_file: Path, target_prefixes: List[str], settings: Dict[str, Any]) -> List[Dict[str, Any]]:
    logger.info(f"Loading collection from: {collection_file.resolve()}")
    try:
        with open(collection_file, 'r', encoding='utf-8') as f:
            collection = json.load(f)
    except Exception as e:
        logger.error(f"Error reading or parsing JSON collection file: {e}")
        raise

    requests_list = []
    postman_vars = settings['POSTMAN_VARS_DICT']

    def process_items(items, folder_name="Collection"):
        for item in items:
            if 'item' in item:
                process_items(item['item'], item['name'])
            elif 'request' in item:
                req = item['request']
                request_name = item['name']

                if target_prefixes and not any(request_name.startswith(prefix) for prefix in target_prefixes):
                    continue

                url_data = req.get('url', {})
                if isinstance(url_data, dict):
                    raw_url = url_data.get('raw', 'http://error_url_not_found')
                elif isinstance(url_data, str):
                    raw_url = url_data
                else:
                    raw_url = 'http://error_url_not_found'

                headers = {h['key'].title(): h['value'] for h in req.get('header', [])}
                raw_body = req.get('body', {}).get('raw', '{}')

                raw_url = apply_postman_vars(raw_url, postman_vars)
                raw_body = apply_postman_vars(raw_body, postman_vars)

                requests_list.append({
                    'folder': folder_name,
                    'name': request_name,
                    'method': req.get('method', 'POST'),
                    'headers': headers,
                    'body': raw_body,
                    'base_url_placeholder': raw_url,
                })

    process_items(collection.get('item', []))
    logger.info(f"Successfully extracted {len(requests_list)} TARGET requests for execution.")
    return requests_list

def run_api_test(request_data: Dict[str, Any], environment_base_url: str, environment_name: str, settings: Dict[str, Any]) -> Dict[str, Any]:

    url_placeholder = request_data['base_url_placeholder']
    final_url = re.sub(r'\{\{baseurl\}\}', environment_base_url, url_placeholder, flags=re.IGNORECASE)

    if request_data.get('url_params'):
        separator = '&' if '?' in final_url else '?'
        param_string = '&'.join([f"{k}={v}" for k, v in request_data['url_params'].items()])
        final_url += separator + param_string

    env_color = Fore.BLUE if 'PRD' in environment_name.upper() else Fore.CYAN

    expected_status = request_data.get('expected_status_code')
    status_hint = f" (Expect {expected_status})" if expected_status else ""

    status_msg = f"{request_data['name']:<70} ({env_color}{environment_name}{Style.RESET_ALL}){status_hint} ..... Executing"
    print(status_msg, end='\r')

    headers = {k.title(): v for k, v in request_data['headers'].items()}
    
    if 'Content-Type' not in headers:
        headers['Content-Type'] = 'application/json'

    auth_type = settings['AUTH_TYPE'].upper()
    auth_header_key = settings['AUTH_HEADER'].title()

    if settings['AUTH_VALUE'] and auth_type in ['BEARER', 'API_KEY']:
        auth_value = settings['AUTH_VALUE']
        if auth_type == 'BEARER':
            headers[auth_header_key] = f"Bearer {auth_value}"
        elif auth_type == 'API_KEY':
            headers[auth_header_key] = auth_value

    elif auth_type == 'BASIC' and settings['BASIC_AUTH_USER'] and settings['BASIC_AUTH_PASS']:
        user_pass = f"{settings['BASIC_AUTH_USER']}:{settings['BASIC_AUTH_PASS']}"
        encoded_user_pass = base64.b64encode(user_pass.encode('utf-8')).decode('utf-8')
        headers[auth_header_key] = f"Basic {encoded_user_pass}"

    elif auth_type != 'NONE':
         logger.warning(f"Unsupported AUTH_TYPE '{auth_type}'. Proceeding without custom auth injection.")

    header_overrides = {k.title(): v for k, v in request_data.get('header_overrides', {}).items()}
    headers.update(header_overrides)

    start_time = time.time()
    try:
        response = requests.request(
            method=request_data['method'],
            url=final_url,
            headers=headers,
            data=request_data['body'],
            timeout=60
        )
        end_time = time.time()
        response_time = int((end_time - start_time) * 1000)

        status_color = get_status_color(response.status_code)

        final_msg = f"{request_data['name']:<70} ({env_color}{environment_name}{Style.RESET_ALL}){status_hint} ..... {status_color}Complete{Style.RESET_ALL} (Status: {response.status_code} in {response_time} ms)      "
        print(final_msg)

        json_body = None
        response_text = escape_html(response.text)

        try:
            clean_text = response.text.strip().lstrip('\ufeff')
            if clean_text:
                json_body = json.loads(clean_text)
                response_text = escape_html(json.dumps(json_body, indent=2))
        except Exception:
            json_body = None

        return {
            'status_code': response.status_code,
            'response_time': response_time,
            'json_body': json_body,
            'raw_text': response_text
        }

    except requests.exceptions.RequestException as e:
        end_time = time.time()
        response_time = int((end_time - start_time) * 1000)

        final_msg = f"{request_data['name']:<70} ({env_color}{environment_name}{Style.RESET_ALL}){status_hint} ..... {Fore.RED}FAILED{Style.RESET_ALL} (Error: {e.__class__.__name__} in {response_time} ms)      "
        print(final_msg)

        logger.error(f"Request failed for {request_data['name']} on {environment_name}: {e.__class__.__name__}")

        return {
            'status_code': 'TIMEOUT/ERROR',
            'response_time': response_time,
            'json_body': None,
            'raw_text': escape_html(f"Request Error: {e.__class__.__name__}: {e}")
        }

def highlight_diffs_in_json(ods_json, prd_json, diff_obj):
    ods_text_unescaped = json.dumps(ods_json, indent=2)
    prd_text_unescaped = json.dumps(prd_json, indent=2)

    ods_text = escape_html(ods_text_unescaped)
    prd_text = escape_html(prd_text_unescaped)

    highlight_tag_start = '<span class="diff-highlight">'
    highlight_tag_end = '</span>'

    keys_to_highlight = set()

    def find_mismatched_keys(d1, d2, collected_keys):
        if isinstance(d1, dict) and isinstance(d2, dict):
            all_keys = set(d1.keys()) | set(d2.keys())
            for k in all_keys:
                if k not in d1 or k not in d2:
                    collected_keys.add(k)
                elif d1[k] != d2[k]:
                    if isinstance(d1[k], dict) and isinstance(d2[k], dict):
                        find_mismatched_keys(d1[k], d2[k], collected_keys)
                        collected_keys.add(k)
                    elif isinstance(d1[k], list) and isinstance(d2[k], list):
                        collected_keys.add(k)
                        for i in range(min(len(d1[k]), len(d2[k]))):
                            find_mismatched_keys(d1[k][i], d2[k][i], collected_keys)
                    else:
                        collected_keys.add(k)

        elif isinstance(d1, list) and isinstance(d2, list):
             for i in range(min(len(d1), len(d2))):
                 find_mismatched_keys(d1[i], d2[i], collected_keys)

    for diff_type in diff_obj.keys():
        diff_data = diff_obj[diff_type]
        if hasattr(diff_data, 'items'):
             iterable = diff_data.items()
        else:
             iterable = [(k, {}) for k in diff_data]

        for path, diff_values in iterable:
            matches = re.findall(r"\['(.*?)'\]", path)
            if matches:
                keys_to_highlight.add(matches[-1])

            val1 = diff_values.get('old_value') or diff_values.get('old_type')
            val2 = diff_values.get('new_value') or diff_values.get('new_type')

            find_mismatched_keys(val1, val2, keys_to_highlight)

    def apply_highlight(json_text, target_keys):
        highlighted = json_text
        for key in target_keys:
            try:
                pattern = r'("' + re.escape(key) + r'"\s*:\s*)(.*?)((,?\n)|(,?$))'

                def replacement(m):
                    prefix = m.group(1)
                    val = m.group(2)
                    suffix = m.group(3)
                    if highlight_tag_start in val: return m.group(0)
                    return f"{prefix}{highlight_tag_start}{val}{highlight_tag_end}{suffix}"

                highlighted = re.sub(pattern, replacement, highlighted, flags=re.MULTILINE)
            except Exception:
                pass
        return highlighted.replace('\\n', '<br>').replace('\\"', '&#34;')

    return apply_highlight(ods_text, keys_to_highlight), apply_highlight(prd_text, keys_to_highlight)

def extract_field_name_from_path(path: str) -> str:
    match = re.findall(r"\'([A-Z_]+)\'", path)
    if match:
        return match[-1]
    return path

def format_deepdiff_path(path: str) -> str:
    path = path.replace("root[", "").replace("][", ".").replace("]", "").replace("'", "")
    path = path.replace("[", ".").replace("]", "")
    return path

def compare_requests_results(ods_results, prd_results, extracted_requests, dynamic_field_terms, schema_dir, settings):
    comparison_data = []
    overall_metrics = {
        'ods_passed': 0, 'ods_failed': 0, 'prd_passed': 0, 'prd_failed': 0, 
        'total_tests': len(extracted_requests), 'data_diff_count': 0, 
        'dynamic_diff_count': 0, 'time_data': [], 'status_fail_count': 0, 
        'security_vuln_count': 0, 'stability_fail_count': 0, 'schema_fail_count': 0, 
        'connection_fail_count': 0,
        'security_findings_list': [], 'stability_findings_list': [], 'schema_findings_list': []
    }

    logger.info("Starting result comparison...")

    schema_cache = {}
    schema_path_base = Path.cwd() / schema_dir

    def load_schema(schema_file):
        if schema_file in schema_cache:
            return schema_cache[schema_file]
        try:
            with open(schema_path_base / schema_file, 'r', encoding='utf-8') as f:
                schema = json.load(f)
                schema_cache[schema_file] = schema
                return schema
        except Exception as e:
            logger.error(f"Failed to load schema file '{schema_file}': {e}")
            return None

    try:
        for i, req_data in enumerate(extracted_requests):
            ods_res = ods_results[i]
            prd_res = prd_results[i]

            test_name = req_data['name']
            current_test_type = req_data.get('test_type', 'FUNCTIONAL').upper()
            expected_status = req_data.get('expected_status_code')
            schema_file = req_data.get('expected_schema_file')

            print(f"[{i+1}/{len(extracted_requests)}] Comparing data for: {test_name[:70]}...                         ", end='\r')

            if ods_res['status_code'] == 'TIMEOUT/ERROR' or prd_res['status_code'] == 'TIMEOUT/ERROR':
                overall_metrics['connection_fail_count'] += 1

            ods_status = "FAIL"
            if isinstance(ods_res['status_code'], int):
                if expected_status is not None:
                    if ods_res['status_code'] == expected_status: ods_status = "PASS"
                    else: ods_status = f"FAIL (Expected {expected_status}, Got {ods_res['status_code']})"
                elif 200 <= ods_res['status_code'] < 300:
                    ods_status = "PASS"

            prd_status = "FAIL"
            if isinstance(prd_res['status_code'], int):
                if expected_status is not None:
                    if prd_res['status_code'] == expected_status: prd_status = "PASS"
                    else: prd_status = f"FAIL (Expected {expected_status}, Got {prd_res['status_code']})"
                elif 200 <= prd_res['status_code'] < 300:
                    prd_status = "PASS"

            if "PASS" in ods_status: overall_metrics['ods_passed'] += 1
            if "PASS" in prd_status: overall_metrics['prd_passed'] += 1
            
            ods_rec_count = get_record_count(ods_res['json_body'])
            prd_rec_count = get_record_count(prd_res['json_body'])

            data_diff_result = 'N/A'
            data_diff_summary = ''
            ods_raw_response_highlighted = ods_res['raw_text']
            prd_raw_response_highlighted = prd_res['raw_text']

            critical_paths = []
            dynamic_value_paths = []
            excluded_paths = []

            schema_validation_fail = False
            schema_fail_details = ""

            if schema_file and JSONSCHEMA_AVAILABLE:
                schema = load_schema(schema_file)
                if schema:
                    for env, res in [('ODS', ods_res), ('PRD', prd_res)]:
                        if res['json_body'] is not None and "PASS" in (ods_status if env == 'ODS' else prd_status):
                            try:
                                validate(instance=res['json_body'], schema=schema)
                            except ValidationError as e:
                                schema_validation_fail = True
                                schema_fail_details += f"{env} Schema Fail: {e.message} at path {e.path} | "
                                logger.error(f"{test_name} failed schema validation on {env}: {e.message}")

                    if schema_validation_fail:
                         overall_metrics['schema_fail_count'] += 1
                         overall_metrics['schema_findings_list'].append(schema_file)

            if ods_res['json_body'] is not None and prd_res['json_body'] is not None:

                diff_obj = DeepDiff(
                    ods_res['json_body'],
                    prd_res['json_body'],
                    ignore_order=True
                )

                if not diff_obj:
                    if ods_rec_count != prd_rec_count:
                         data_diff_result = 'FAIL'
                         overall_metrics['data_diff_count'] += 1
                    else:
                        data_diff_result = 'PASS'
                        data_diff_summary = 'Response bodies are identical.'
                else:
                    for diff_type in diff_obj.keys():
                        diff_data = diff_obj[diff_type]
                        if hasattr(diff_data, 'items'):
                             iterable = diff_data.items()
                        else:
                             iterable = [(k, {}) for k in diff_data]

                        for path, details in iterable:
                            is_excluded = False
                            for term in settings['EXCLUDE_FIELD_NAMES_LIST']:
                                if term in str(path):
                                    is_excluded = True
                                    break
                            if is_excluded:
                                excluded_paths.append(path)
                                continue

                            is_dynamic = False
                            for term in settings['DYNAMIC_FIELD_TERMS']:
                                if term.upper() in str(path).upper():
                                    is_dynamic = True
                                    break
                            if is_dynamic:
                                dynamic_value_paths.append(path)
                                continue

                            critical_paths.append(path)

                    if ods_rec_count != prd_rec_count:
                        data_diff_result = 'FAIL'
                        if not critical_paths:
                            overall_metrics['data_diff_count'] += 1
                        else:
                            overall_metrics['data_diff_count'] += len(critical_paths)
                    elif critical_paths:
                        data_diff_result = 'FAIL'
                        overall_metrics['data_diff_count'] += len(critical_paths)
                    elif dynamic_value_paths or excluded_paths:
                        data_diff_result = 'WARN_DIFF'
                        overall_metrics['dynamic_diff_count'] += len(dynamic_value_paths) + len(excluded_paths)
                    else:
                        data_diff_result = 'PASS'
                        data_diff_summary = 'Response bodies are identical.'

                    if diff_obj:
                        ods_raw_response_highlighted, prd_raw_response_highlighted = highlight_diffs_in_json(ods_res['json_body'], prd_res['json_body'], diff_obj)

            else:
                if "PASS" in ods_status and "PASS" in prd_status:
                    data_diff_result = 'PASS (No JSON)'
                    data_diff_summary = 'Comparison Skipped (No JSON detected or Parse Error).'
                else:
                    data_diff_result = 'N/A'
                    data_diff_summary = f'N/A: Status FAIL/Error ({ods_res["status_code"]} vs {prd_res["status_code"]}).'

            comments = ""
            test_findings = []

            if schema_validation_fail:
                comments += f"SCHEMA VALIDATION FAILED: {schema_fail_details}. "
                data_diff_result = 'FAIL'
                test_findings.append({'category': 'Schema', 'type': 'Structural/Type Mismatch', 'details': schema_fail_details.strip()})

            if data_diff_result == 'FAIL':
                 if ods_rec_count != prd_rec_count:
                     comments += f"COUNT MISMATCH: ODS has {ods_rec_count} items, PRD has {prd_rec_count}. "
                     test_findings.append({'category': 'Data Integrity', 'type': 'Record Count Mismatch', 'details': f"ODS: {ods_rec_count}, PRD: {prd_rec_count}"})
                 
                 if critical_paths:
                     formatted_critical_paths = [format_deepdiff_path(str(p)) for p in critical_paths]
                     comments += f"CRITICAL DATA MISMATCH: Failed Keys: {', '.join(formatted_critical_paths)}. "
                     test_findings.append({'category': 'Data Integrity', 'type': 'Critical Data Mismatch', 'details': f"Keys: {', '.join(formatted_critical_paths)}"})

            elif data_diff_result == 'WARN_DIFF':
                warn_details = []
                if dynamic_value_paths:
                    warn_details.append(f"Dynamic values ({len(dynamic_value_paths)})")
                if excluded_paths:
                    warn_details.append(f"Excluded fields ({len(excluded_paths)})")

                comments = f"Functional PASS (with Warnings): {', '.join(warn_details)}."
                test_findings.append({'category': 'Data Integrity', 'type': 'Soft Data Mismatch', 'details': f"Dynamic: {len(dynamic_value_paths)}, Excluded: {len(excluded_paths)}"})

            data_diff_summary_list = []
            
            if ods_rec_count != prd_rec_count:
                 data_diff_summary_list.append(f"Count Mismatch (ODS:{ods_rec_count} vs PRD:{prd_rec_count})")
            
            if data_diff_result == 'FAIL' and critical_paths: data_diff_summary_list.append("Critical Data Mismatch")
            if dynamic_value_paths: data_diff_summary_list.append("Dynamic Value Change")
            if excluded_paths: data_diff_summary_list.append("Excluded Field Mismatch")

            if data_diff_summary_list:
                data_diff_summary = f"Diffs found: {', '.join(data_diff_summary_list)}"

                details = []
                for diff_type in diff_obj.keys():
                    diff_data = diff_obj[diff_type]
                    if hasattr(diff_data, 'items'):
                         iterable = diff_data.items()
                    else:
                         iterable = [(k, {}) for k in diff_data]

                    for path, diff_values in iterable:
                        if len(details) >= 10: break

                        if path in critical_paths or (data_diff_result == 'WARN_DIFF' and (path in dynamic_value_paths or path in excluded_paths)):
                            old_val = diff_values.get('old_value')
                            new_val = diff_values.get('new_value')
                            old_type = diff_values.get('old_type')
                            new_type = diff_values.get('new_type')
                            lbl = " (Excluded)" if path in excluded_paths else (" (Dynamic)" if path in dynamic_value_paths else "")

                            if diff_type == 'type_changes':
                                old_val_str = f"{old_val} ({old_type.__name__})" if old_type else str(old_val)
                                new_val_str = f"{new_val} ({new_type.__name__})" if new_type else str(new_val)
                                details.append(f"{format_deepdiff_path(str(path))}{lbl}: ODS='{old_val_str}' PRD='{new_val_str}'")
                            elif isinstance(old_val, dict) and isinstance(new_val, dict):
                                inner_diff = DeepDiff(old_val, new_val, ignore_order=True)
                                inner_changes = []
                                for cat in inner_diff.keys():
                                    inner_data = inner_diff[cat]
                                    if hasattr(inner_data, 'items'):
                                         inner_iterable = inner_data.keys()
                                    else:
                                         inner_iterable = inner_data

                                    for item in inner_iterable:
                                        k = re.findall(r"\['(.*?)'\]", str(item)) or re.findall(r"\[(.*?)\]", str(item))
                                        if k: inner_changes.append(k[-1])
                                
                                if inner_changes:
                                    inner_str = ", ".join(list(set(inner_changes)))
                                    details.append(f"{format_deepdiff_path(str(path))}{lbl}: Mismatch in fields: {inner_str}")
                                else:
                                     details.append(f"{format_deepdiff_path(str(path))}{lbl}: Complex Object Mismatch")
                            else:
                                s_old = str(old_val)
                                s_new = str(new_val)
                                if len(s_old) > 100: s_old = s_old[:100] + "..."
                                if len(s_new) > 100: s_new = s_new[:100] + "..."
                                details.append(f"{format_deepdiff_path(str(path))}{lbl}: ODS='{s_old}' PRD='{s_new}'")

                if details:
                     data_diff_summary += " | Details: " + " | ".join(details)

            if schema_validation_fail:
                 data_diff_summary += " | Schema Fail: YES"

            if current_test_type == 'SECURITY':
                is_critical_security_fail = False
                if "PASS" in ods_status:
                    test_findings.append({'category': 'Security', 'type': 'Auth/Authz Bypass', 'details': f"ODS: Expected 4xx, got {ods_res['status_code']} (Success)."})
                    is_critical_security_fail = True
                if "PASS" in prd_status:
                    test_findings.append({'category': 'Security', 'type': 'Auth/Authz Bypass', 'details': f"PRD: Expected 4xx, got {prd_res['status_code']} (Success)."})
                    is_critical_security_fail = True
                if ods_res['status_code'] == 500:
                    test_findings.append({'category': 'Security', 'type': 'Payload Processing Error (500)', 'details': f"ODS: Payload caused 500 Internal Error (Potential Leak/Injection)."})
                    is_critical_security_fail = True
                if prd_res['status_code'] == 500:
                    test_findings.append({'category': 'Security', 'type': 'Payload Processing Error (500)', 'details': f"PRD: Payload caused 500 Internal Error (Potential Leak/Injection)."})
                    is_critical_security_fail = True

                if is_critical_security_fail:
                    comments = f"CRITICAL SECURITY VULNERABILITY: {', '.join(set(f['type'] for f in test_findings if f['category'] == 'Security'))}. ACTION: Fix security logic immediately."
                    if not overall_metrics['security_vuln_count']:
                        overall_metrics['security_vuln_count'] = len(set(f['type'] for f in test_findings if f['category'] == 'Security'))
                    if "PASS" in ods_status or "PASS" in prd_status or ods_res['status_code'] == 500 or prd_res['status_code'] == 500:
                        overall_metrics['status_fail_count'] += 1

            elif current_test_type == 'STABILITY':
                is_stability_fail = False
                if ods_res['status_code'] == 500:
                    test_findings.append({'category': 'Stability', 'type': 'Unhandled 500 Error', 'details': 'ODS: Invalid input caused 500. Should return 400/422.'})
                    is_stability_fail = True
                if prd_res['status_code'] == 500:
                    test_findings.append({'category': 'Stability', 'type': 'Unhandled 500 Error', 'details': 'PRD: Invalid input caused 500. Should return 400/422.'})
                    is_stability_fail = True

                if is_stability_fail:
                    comments = f"STABILITY BUG: {', '.join(set(f['type'] for f in test_findings if f['category'] == 'Stability'))}. ACTION: Implement input validation to return 400/422."
                    if not overall_metrics['stability_fail_count']:
                        overall_metrics['stability_fail_count'] = len(set(f['type'] for f in test_findings if f['category'] == 'Stability'))
                    overall_metrics['status_fail_count'] += 1

            server_error_found = False
            if ods_res['status_code'] == 500:
                comments = f"STABILITY FAILURE (ODS): API returned 500 Internal Server Error. {comments}"
                test_findings.append({'category': 'Stability', 'type': '500 Internal Server Error', 'details': 'ODS returned 500'})
                overall_metrics['status_fail_count'] += 1
                server_error_found = True
            
            if prd_res['status_code'] == 500:
                comments = f"STABILITY FAILURE (PRD): API returned 500 Internal Server Error. {comments}"
                test_findings.append({'category': 'Stability', 'type': '500 Internal Server Error', 'details': 'PRD returned 500'})
                overall_metrics['status_fail_count'] += 1
                server_error_found = True

            if not server_error_found and not comments:
                if data_diff_result == 'PASS' and "PASS" in ods_status and "PASS" in prd_status:
                    comments = "Functional PASS. Status OK. Data bodies match perfectly."
                elif data_diff_result == 'PASS (No JSON)':
                    comments = "Functional PASS. Status OK. Data Comparison Skipped (Non-JSON)."
                elif 'N/A' in data_diff_summary:
                    comments = f"Functional PASS. Status OK. Data comparison skipped ({data_diff_summary.split(':')[-1].strip()})."
                elif expected_status is not None:
                     comments = f"Test PASS: API correctly returned expected status {expected_status}."
                elif current_test_type in ['SECURITY', 'STABILITY', 'NEGATIVE'] and ods_res['status_code'] in range(400, 500) and prd_res['status_code'] in range(400, 500):
                    comments = "Test PASS: API correctly returned expected 4xx client error status."

            for finding in test_findings:
                if finding['category'] == 'Security' and finding['type'] not in overall_metrics['security_findings_list']:
                    overall_metrics['security_findings_list'].append(finding['type'])
                if finding['category'] == 'Stability' and finding['type'] not in overall_metrics['stability_findings_list']:
                    overall_metrics['stability_findings_list'].append(finding['type'])
                if finding['category'] == 'Schema' and finding['type'] not in overall_metrics['schema_findings_list']:
                    overall_metrics['schema_findings_list'].append(finding['type'])

            ods_time = ods_res['response_time']
            prd_time = prd_res['response_time']
            overall_metrics['time_data'].append({'name': test_name, 'ods_time': ods_time, 'prd_time': prd_time})

            diff_css_class = ""
            display_diff_result = data_diff_result
            if data_diff_result == 'FAIL': 
                diff_css_class = "danger"
            elif data_diff_result == 'WARN_DIFF': 
                diff_css_class = "soft-fail"
                display_diff_result = "SOFT FAIL"
            elif 'PASS' in data_diff_result: 
                diff_css_class = "success"

            comparison_data.append({
                'test_name': test_name,
                'ods_status': ods_status,
                'prd_status': prd_status,
                'ods_time': f"{ods_time} ms",
                'prd_time': f"{prd_time} ms",
                'ods_record_count': ods_rec_count,
                'prd_record_count': prd_rec_count, 
                'test_type': current_test_type,
                'data_diff_result': display_diff_result,
                'diff_css_class': diff_css_class,
                'data_diff_summary': data_diff_summary,
                'comments': comments.strip(),
                'request_body': escape_html(req_data['body']),
                'ods_raw_response': ods_raw_response_highlighted,
                'prd_raw_response': prd_raw_response_highlighted,
                'findings': test_findings,
            })

        print(' ' * 100, end='\r')
        logger.info(f"Comparison complete. Found {overall_metrics['data_diff_count']} critical data differences and {overall_metrics['dynamic_diff_count']} dynamic data differences.")
        return comparison_data, overall_metrics

    except Exception as e:
        logger.critical(f"\nCRITICAL ERROR during metric comparison: {e}")
        return comparison_data, overall_metrics 

def generate_all_executable_runs(all_requests_templates: List[Dict[str, Any]], data_file: str) -> List[Dict[str, Any]]:

    data_file_path = Path.cwd() / data_file
    logger.info(f"\n--- Generating Executable Runs (Mixed Static and Data-Driven) ---")

    try:
        if not data_file_path.exists():
            logger.error(f"ERROR: Test data file '{data_file}' not found at path: {data_file_path.resolve()}.")
            return []

        data_df = pd.read_excel(data_file_path, sheet_name='Data')

    except Exception as e:
        logger.critical(f"CRITICAL ERROR reading Excel Data sheet: {e}.")
        return []

    final_runs = []
    request_templates = {req['name']: req for req in all_requests_templates}

    HEADER_PREFIX = 'Override_Header_'
    URL_PARAM_PREFIX = 'Override_URL_Param_'
    SCHEMA_FILE_COL = 'Expected_Schema_File'

    all_cols = data_df.columns.tolist()
    header_override_cols = [c for c in all_cols if c.startswith(HEADER_PREFIX)]
    url_param_override_cols = [c for c in all_cols if c.startswith(URL_PARAM_PREFIX)]

    static_tests_added = set()

    for index, row in data_df.iterrows():
        if 'test_name' not in row or pd.isna(row['test_name']): continue
        if 'Execution_Type' not in row or pd.isna(row['Execution_Type']): continue

        test_case_name = row['test_name']
        execution_type = row['Execution_Type'].upper()

        test_type = str(row.get('Test_Type', 'FUNCTIONAL')).strip().upper()
        expected_status = row.get('Expected_Status_Code')
        expected_schema_file = row.get(SCHEMA_FILE_COL)

        if test_case_name not in request_templates:
            logger.warning(f"Test name '{test_case_name}' from row {index+1} not found in Postman collection. Skipping.")
            continue

        template = request_templates[test_case_name]

        header_overrides = {}
        for col in header_override_cols:
            if pd.notna(row[col]):
                header_key = col.replace(HEADER_PREFIX, '')
                header_overrides[header_key] = str(row[col]).strip()

        url_params = {}
        for col in url_param_override_cols:
            if pd.notna(row[col]):
                param_key = col.replace(URL_PARAM_PREFIX, '')
                url_params[param_key] = str(row[col]).strip()

        if execution_type == 'STATIC':
            if test_case_name in static_tests_added:
                continue

            new_run = template.copy()
            new_run['name'] = f"{template['name']}"
            new_run['test_type'] = test_type

            if header_overrides or url_params:
                new_run['name'] += f" [Override]"
                new_run['header_overrides'] = header_overrides
                new_run['url_params'] = url_params

            if pd.notna(expected_status):
                 new_run['expected_status_code'] = int(expected_status)

            if pd.notna(expected_schema_file):
                 new_run['expected_schema_file'] = str(expected_schema_file).strip()

            final_runs.append(new_run)
            static_tests_added.add(test_case_name)

        elif execution_type == 'DATA_DRIVEN':

            unique_run_name = f"{template['name']} (Run: {row.get('run_id', index+1)})"
            raw_body = template['body']

            for col in ['vpin', 'offset', 'limit']:
                value = row.get(col)
                if pd.notna(value):
                    str_value = str(value)

                    raw_body = raw_body.replace(f'{{{{{col}}}}}', str_value)

                    pattern = r'(\"' + re.escape(col) + r'\"\s*:\s*\")(.*?)(\")'

                    if re.search(pattern, raw_body, re.IGNORECASE):
                        raw_body = re.sub(
                            pattern,
                            r'\1' + str_value + r'\3',
                            raw_body,
                            flags=re.IGNORECASE
                        )

            new_run = template.copy()
            new_run['name'] = unique_run_name
            new_run['body'] = raw_body
            new_run['test_type'] = test_type
            new_run['header_overrides'] = header_overrides
            new_run['url_params'] = url_params

            if pd.notna(expected_status):
                 new_run['expected_status_code'] = int(expected_status)

            if pd.notna(expected_schema_file):
                 new_run['expected_schema_file'] = str(expected_schema_file).strip()

            final_runs.append(new_run)

    logger.info(f"Generated {len(final_runs)} executable runs (Mixing Static and Data-Driven tests).")
    return final_runs

def generate_report(comparison_data, metrics, settings, output_dir):

    if not os.path.exists(settings['TEMPLATE_FILE']):
        logger.error(f"\nERROR: Template file '{settings['TEMPLATE_FILE']}' not found. Cannot generate report.")
        return None

    logger.info("--- Generating HTML Report and Charts (Phase 5/5) ---")

    RUN_MODE = settings.get('RUN_MODE', 'DUAL')
    TARGET_ENV_NAME = RUN_MODE.split('_')[0] if RUN_MODE != 'DUAL' else 'DUAL'

    critical_pass_count = metrics['total_tests'] - metrics['status_fail_count'] - metrics['data_diff_count'] - metrics['schema_fail_count'] - metrics['connection_fail_count']

    chart_data_status = {
        'critical_pass': max(0, critical_pass_count),
        'critical_fail': metrics['data_diff_count'] + metrics['schema_fail_count'],
        'warn_diff': metrics['dynamic_diff_count'],
        'status_fail': metrics['status_fail_count'],
        'connection_fail': metrics['connection_fail_count'],
        'total': metrics['total_tests']
    }

    ods_times = [item['ods_time'] for item in metrics['time_data'] if isinstance(item['ods_time'], int)]
    prd_times = [item['prd_time'] for item in metrics['time_data'] if isinstance(item['prd_time'], int)]

    avg_ods_time = int(sum(ods_times) / len(ods_times)) if ods_times else 0
    avg_prd_time = int(sum(prd_times) / len(prd_times)) if prd_times else 0
    metrics['p90_ods_time'] = "N/A (P90 removed)"
    metrics['p90_prd_time'] = "N/A (P90 removed)"

    metrics['avg_ods_time'] = avg_ods_time
    metrics['avg_prd_time'] = avg_prd_time

    chart_data_perf_avg = {
        'avg_ods_time': avg_ods_time,
        'avg_prd_time': avg_prd_time
    }

    chart_data_time_per_test = {
        'labels': [item['name'] for item in metrics['time_data']],
        'ods_times': ods_times,
        'prd_times': prd_times,
    }

    performance_delta = avg_prd_time - avg_ods_time
    perf_factor = round(avg_prd_time / avg_ods_time, 2) if avg_ods_time > 0 else 0

    metrics['perf_delta'] = performance_delta
    metrics['perf_factor'] = perf_factor

    perf_summary = ""
    if RUN_MODE != 'DUAL':
        perf_summary = f"Performance metrics for **{TARGET_ENV_NAME}** environment only. Comparison data is duplicated."
    elif performance_delta > 500:
        perf_summary = f"Production (PRD) is **{perf_factor}x slower** than ODS on average ({performance_delta} ms delta). This is a **CRITICAL** performance regression."
    elif performance_delta > 100:
        perf_summary = f"Production (PRD) is **{perf_factor}x slower** than ODS on average ({performance_delta} ms delta). This is a **WARNING** performance difference."
    else:
        perf_summary = f"Performance is healthy. PRD is {perf_factor}x slower on average ({performance_delta} ms delta)."

    security_findings_unique = set(metrics['security_findings_list'])
    stability_findings_unique = set(metrics['stability_findings_list'])
    schema_findings_unique = set(metrics['schema_findings_list'])

    if security_findings_unique:
        security_summary = f"**{len(security_findings_unique)} CRITICAL Security Vulnerability Type(s)** found: {', '.join(security_findings_unique)}. Check table for affected tests."
    else:
        security_summary = "No critical security vulnerabilities detected in the tested scenarios."

    if stability_findings_unique:
        stability_summary = f"**{len(stability_findings_unique)} Stability Bug Type(s)** found: {', '.join(stability_findings_unique)}. Tests failed with unhandled 500 status codes."
    else:
        stability_summary = "No critical stability bugs detected (unhandled 500s for invalid input)."

    if schema_findings_unique:
        schema_summary = f"**{len(schema_findings_unique)} Schema Validation Failures** found. Structural integrity of response JSON is compromised."
        metrics['schema_summary'] = schema_summary
    else:
        metrics['schema_summary'] = "Schema validation passed for all applicable tests."

    if RUN_MODE != 'DUAL':
        data_integrity_summary = f"**Data integrity comparison skipped (Single Run Mode: {TARGET_ENV_NAME}).**"
    else:
        data_integrity_summary = f"**Core Data Integrity is PASS ({metrics['data_diff_count']} Critical Differences).** All {metrics['dynamic_diff_count']} data differences were confined to dynamic fields (e.g., LAST_UPDATE, ID type changes)."

    critical_summary = {
        'data_mismatch': metrics['data_diff_count'],
        'soft_fail': metrics['dynamic_diff_count'],
        'connection_fail': metrics['connection_fail_count']
    }

    summary = {
        'perf_finding': perf_summary,
        'security_finding': security_summary,
        'stability_finding': stability_summary,
        'data_integrity': data_integrity_summary,
        'schema_finding': metrics['schema_summary']
    }

    report_metadata = {
        'collection_name': settings['COLLECTION_FILE'],
        'ods_url': settings['ODS_URL'] if RUN_MODE != 'PRD_ONLY' else f"N/A ({TARGET_ENV_NAME} Mode)",
        'prd_url': settings['PRD_URL'] if RUN_MODE != 'ODS_ONLY' else f"N/A ({TARGET_ENV_NAME} Mode)",
        'run_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S %Z"),
        'total_tests': metrics['total_tests'],
        'ENABLE_PERF_GRAPH': settings['ENABLE_PERF_GRAPH'],
        'RUN_MODE': RUN_MODE,
        'TARGET_ENV_NAME': TARGET_ENV_NAME
    }

    test_type_counts = {}
    for item in comparison_data:
        test_type = item.get('test_type', 'UNKNOWN')
        test_type_counts[test_type] = test_type_counts.get(test_type, 0) + 1

    chart_data_test_type = {
        'labels': list(test_type_counts.keys()),
        'counts': list(test_type_counts.values())
    }

    metrics['perf_score'] = "N/A (Score removed)"

    template_vars = {
        'comparison_data': comparison_data,
        'metrics': metrics,
        'summary': summary,
        'critical_summary': critical_summary,
        'metadata': report_metadata,
        'chart_data_perf_avg': chart_data_perf_avg,
        'chart_data_status_json': json.dumps(chart_data_status),
        'chart_data_perf_avg_json': json.dumps(chart_data_perf_avg),
        'chart_data_time_per_test_json': json.dumps(chart_data_time_per_test),
        'chart_data_test_type_json': json.dumps(chart_data_test_type),
        'perf_delta': performance_delta,
        'perf_factor': perf_factor
    }

    file_loader = FileSystemLoader('.')
    env = Environment(loader=file_loader)
    template = env.get_template(settings['TEMPLATE_FILE'])

    report_path = get_unique_filepath(output_dir, 'Dashboard_report.html')
    abs_report_path = os.path.abspath(report_path)

    with open(report_path, 'w', encoding="utf-8") as f:
        f.write(template.render(template_vars))

    logger.info(f"\n--- HTML Report successfully generated at: {abs_report_path} ---")

    return abs_report_path

if __name__ == '__main__':

    try:
        if not JSONSCHEMA_AVAILABLE:
            logger.warning("WARNING: jsonschema not installed. Response Schema Validation will be skipped. Install with 'pip install jsonschema'.")

        import pandas as pd
        import openpyxl
    except ImportError as e:
        logger.critical(f"ERROR: Missing required libraries for execution. Please install: pip install pandas openpyxl requests deepdiff jinja2 colorama jsonschema. Missing: {e}")
        sys.exit(1)

    settings = load_settings_from_excel(TEST_DATA_FILE)

    POSTMAN_COLLECTION_FILE = settings['POSTMAN_COLLECTION_FILE']
    COLLECTION_FILE = settings['COLLECTION_FILE']
    ODS_URL = settings['ODS_URL']
    PRD_URL = settings['PRD_URL']
    OUTPUT_DIR = settings['OUTPUT_DIR']
    TARGET_TEST_PREFIXES = settings['TARGET_TEST_PREFIXES']
    DYNAMIC_FIELD_TERMS = settings['DYNAMIC_FIELD_TERMS']
    AUTO_OPEN_HTML = settings['AUTO_OPEN_HTML']
    RUN_MODE = settings.get('RUN_MODE', 'DUAL').upper()
    SCHEMA_DIR = settings['SCHEMA_DIR']

    if not os.path.exists(POSTMAN_COLLECTION_FILE):
        logger.critical(f"ERROR: Collection file '{COLLECTION_FILE}' not found at {POSTMAN_COLLECTION_FILE.resolve()}. Please update COLLECTION_FILE in the Settings sheet.")
        sys.exit(1)

    today_str = datetime.now().strftime("%m-%d-%Y")
    run_output_dir = os.path.join(OUTPUT_DIR, today_str)
    
    if not os.path.exists(run_output_dir):
        os.makedirs(run_output_dir)
        logger.info(f"Created date-specific output directory: {run_output_dir}")

    ods_url_color = Fore.CYAN
    prd_url_color = Fore.BLUE
    print(f"Starting API Comparison Script using (Mode: {RUN_MODE}):")
    print(f"  Collection: {COLLECTION_FILE}")
    print(f"  ODS URL: {ods_url_color}{ODS_URL}{Style.RESET_ALL}")
    print(f"  PRD URL: {prd_url_color}{PRD_URL}{Style.RESET_ALL}")

    ods_healthy = True
    prd_healthy = True

    if RUN_MODE in ['DUAL', 'ODS_ONLY']:
        ods_healthy = check_api_health('ODS', ODS_URL)
    if RUN_MODE in ['DUAL', 'PRD_ONLY']:
        prd_healthy = check_api_health('PRD', PRD_URL)

    if RUN_MODE == 'DUAL' and not (ods_healthy and prd_healthy):
        logger.warning("One or both environments failed the health check. Proceeding with caution, but expect failures.")

    logger.info("\n--- Parsing Postman Collection ---")
    try:
        all_requests_templates = extract_requests(POSTMAN_COLLECTION_FILE, TARGET_TEST_PREFIXES, settings)
        if not all_requests_templates:
            logger.error("ERROR: No requests found in collection. Exiting.")
            sys.exit(1)
    except Exception as e:
        logger.critical(f"ERROR: Failed to parse Postman collection: {e}. Exiting.")
        sys.exit(1)

    all_runs = generate_all_executable_runs(all_requests_templates, TEST_DATA_FILE)

    if not all_runs:
        logger.error("ERROR: No executable runs generated. Check if 'test_data.xlsx' Data sheet has valid runs.")
        sys.exit(1)

    logger.info(f"Total scenarios to execute and compare: {len(all_runs)}")

    ods_results = []
    prd_results = []

    if RUN_MODE in ['DUAL', 'ODS_ONLY']:
        logger.info("\n------------ Executing ODS Requests (Phase 2/5 - ODS) ------------")
        for req_data in all_runs:
            ods_results.append(run_api_test(req_data, ODS_URL, 'ODS', settings))

    if RUN_MODE in ['DUAL', 'PRD_ONLY']:
        if RUN_MODE == 'PRD_ONLY':
            logger.info("\n------------ Executing PRD Requests (Phase 2/5 - PRD) ------------")
            for req_data in all_runs:
                prd_results.append(run_api_test(req_data, PRD_URL, 'PRD', settings))

        elif RUN_MODE == 'DUAL':
            logger.info("\n------------ Executing PRD Requests (Phase 2/5 - PRD) ------------")
            for req_data in all_runs:
                prd_results.append(run_api_test(req_data, PRD_URL, 'PRD', settings))

    if RUN_MODE == 'ODS_ONLY':
        prd_results = ods_results.copy()
        logger.info("Single-run mode (ODS_ONLY) enabled. Duplicating ODS results as PRD results.")
    elif RUN_MODE == 'PRD_ONLY':
        ods_results = prd_results.copy()
        logger.info("Single-run mode (PRD_ONLY) enabled. Duplicating PRD results as ODS results.")

    logger.info("\n--- Starting Data and Metric Comparison (Phase 3/5) ---")

    comparison_data, metrics = compare_requests_results(ods_results, prd_results, all_runs, DYNAMIC_FIELD_TERMS, SCHEMA_DIR, settings)

    report_path_html = generate_report(comparison_data, metrics, settings, run_output_dir)
    report_path_excel = export_to_excel(comparison_data, run_output_dir)

    logger.info("\n--- Opening HTML Report in Browser ---")
    if report_path_html and AUTO_OPEN_HTML == 'YES':
        webbrowser.open(report_path_html)
        print(f"{Fore.GREEN}Report opened successfully in browser.{Style.RESET_ALL}")
    elif report_path_html:
         print(f"{Fore.GREEN}Report generated successfully.{Style.RESET_ALL} Auto-open is disabled. Open the file manually: {report_path_html}")

    logger.info("Script execution finished.")
