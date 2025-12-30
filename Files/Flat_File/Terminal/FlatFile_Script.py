import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import numpy as np
import math
import traceback
import sys
import subprocess
from datetime import datetime
import csv
import re
from jinja2 import Environment, FileSystemLoader
import time

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

class Colors:
    HEADER = '\033[95m'; OKBLUE = '\033[94m'; OKCYAN = '\033[96m'; OKGREEN = '\033[92m'
    WARNING = '\033[93m'; FAIL = '\033[91m'; ENDC = '\033[0m'; BOLD = '\033[1m'; UNDERLINE = '\033[4m'

def clear_screen():
    if os.name == 'nt': os.system('cls'); os.system('')
    else: os.system('clear')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, 'Configuration.xlsx')
SQL_RULES_DIR = os.path.join(SCRIPT_DIR, 'sql_rules')

def get_unique_filepath(filepath):
    if not os.path.exists(filepath): return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base}_{counter}{ext}"
        if not os.path.exists(new_filepath): return new_filepath
        counter += 1

def format_and_autofit_sheet(worksheet, df, header_list, write_index=False, enable_autofit=True):
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in worksheet[1]: cell.font = header_font; cell.fill = header_fill
    
    if not enable_autofit:
        for i, col_name in enumerate(header_list):
            worksheet.column_dimensions[get_column_letter(i + 1 + (1 if write_index else 0))].width = len(str(col_name)) + 5
        return

    col_map = {i + 1 + (1 if write_index else 0): col for i, col in enumerate(header_list)}
    if write_index: col_map[1] = df.index.name if df.index.name else "S.No"
    for col_idx, col_name in col_map.items():
        max_len = len(str(col_name))
        try:
            series_pos = col_idx - (1 if write_index else 0) - 1
            if series_pos < 0 or series_pos >= len(df.columns): continue
            
            series = df.index.to_series() if write_index and col_idx == 1 else df.iloc[:, series_pos]
            
            if len(series) > 5000:
                sample_series = series.sample(n=5000, random_state=1).astype(str)
            else:
                sample_series = series.astype(str)
                
            max_len_val = sample_series.map(len).max()
            if pd.notna(max_len_val): max_len = max(max_len, int(max_len_val))
        except Exception: pass
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

def highlight_detail_rows(writer, sheet_name):
    worksheet = writer.sheets.get(sheet_name)
    if not worksheet or worksheet.max_row <= 1: return
    light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = light_red_fill

def highlight_overall_status(writer, status):
    worksheet = writer.sheets.get('Summary')
    if not worksheet: return
    fill_color = None
    if status == 'FAIL': fill_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif status == 'PASS': fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if fill_color: worksheet['B3'].fill = fill_color

def highlight_failed_cells(writer, df, failed_coords_set):
    if df.empty or not failed_coords_set: return
    worksheet = writer.sheets.get('Data_Failed')
    if not worksheet: return
    light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    original_to_new_index = {orig_idx: i for i, orig_idx in enumerate(df.index, start=2)}
    for coord in failed_coords_set:
        original_df_index, _, col_pos = coord
        if original_df_index in original_to_new_index:
            row_idx_in_excel = original_to_new_index[original_df_index]
            col_idx_in_excel = col_pos + 2
            worksheet.cell(row=row_idx_in_excel, column=col_idx_in_excel).fill = light_red_fill

def clean_dateformat(value):
    if pd.isna(value): return np.nan
    val_stripped = str(value).strip()
    if val_stripped in ['nan', 'None', 'NaT', '']: return np.nan
    else: return val_stripped

def generate_error_summary(results):
    summary_data = []
    error_mappings = {
        'Configuration_Issues': 'Configuration',
        'Data_Content': 'Mandatory',
        'DataType': 'Data Type',
        'Date_Format': 'Date Format',
        'Length_Compare': 'Length',
        'Content_Validation': 'Content Rule',
        'WhiteSpace': 'Whitespace',
        'SQL_Exclusions': 'SQL Exclusion'
    }
    for key, name in error_mappings.items():
        df = results.get(key)
        if df is not None and not df.empty and 'Column Name' in df.columns:
            counts = df['Column Name'].value_counts().reset_index()
            counts.columns = ['Column Name', 'Error Count']
            counts['Error Type'] = name
            summary_data.append(counts)
        elif key == 'SQL_Exclusions' and df is not None and not df.empty:
            summary_data.append(pd.DataFrame([{'Column Name': 'N/A', 'Error Count': len(df), 'Error Type': name}]))

    if not summary_data:
        return pd.DataFrame(columns=['Error Type', 'Column Name', 'Error Count'])

    summary_df = pd.concat(summary_data, ignore_index=True)
    return summary_df[['Error Type', 'Column Name', 'Error Count']]

def generate_html_report(tc_id, description, results, final_output_file):
    try:
        env = Environment(loader=FileSystemLoader(SCRIPT_DIR))
        template = env.get_template('report_template.html')

        summary_df = results['Summary']
        summary_dict = dict(zip(summary_df['Metric'], summary_df['Value']))
        
        def get_status_class(val):
            if isinstance(val, str):
                val_lower = val.lower()
                if val_lower == 'pass': return 'status-pass'
                if val_lower == 'fail': return 'status-fail'
                if val_lower == 'warn': return 'status-warn'
            return ''

        total_records_processed = int(summary_dict.get('Total Records (Post-SQL Filter)', 0))
        records_with_errors = int(summary_dict.get('Records with Errors', 0))
        sql_exclusions = len(results.get('SQL_Exclusions', pd.DataFrame()))
        records_fully_passed = total_records_processed - records_with_errors

        pie_labels = ['Records Passed', 'Records with Errors', 'Excluded by SQL Filter']
        pie_data = [records_fully_passed, records_with_errors, sql_exclusions]
        pie_colors = ['#4CAF50', '#F44336', '#FFC107']
        
        error_summary_df = results.get('Error_Summary', pd.DataFrame())
        bar_labels = []
        bar_data = []
        
        if not error_summary_df.empty:
            error_breakdown = error_summary_df.groupby('Error Type')['Error Count'].sum().sort_values(ascending=False)
            bar_labels = error_breakdown.index.tolist()
            bar_data = error_breakdown.values.tolist()
            
        chart_data = {
            'pie_labels': pie_labels,
            'pie_data': pie_data,
            'pie_colors': pie_colors,
            'bar_labels': bar_labels,
            'bar_data': bar_data
        }

        html_sections = {}
        sections_to_render = [
            'Error_Summary', 'Configuration_Issues', 'Structure', 'Column_Comparison', 'Column_Sequence',
            'WhiteSpace', 'Data_Content', 'DataType', 'Date_Format', 'Length_Compare',
            'Content_Validation', 'Data_Failed', 'Duplicate', 'SQL_Exclusions'
        ]

        inherently_fail_sections = {
            'Error_Summary', 'Configuration_Issues', 'WhiteSpace', 'Data_Content', 
            'DataType', 'Date_Format', 'Length_Compare', 'Content_Validation', 
            'Data_Failed', 'Duplicate', 'SQL_Exclusions'
        }

        for section_name in sections_to_render:
            df = results.get(section_name)
            if df is not None and not df.empty:
                section_status = 'status-neutral'
                if section_name in inherently_fail_sections:
                    section_status = 'status-fail'
                elif 'Status' in df.columns:
                    if 'FAIL' in df['Status'].unique():
                        section_status = 'status-fail'
                    else:
                        section_status = 'status-pass'

                styler = df.style
                if 'Status' in df.columns:
                    class_df = pd.DataFrame('', index=df.index, columns=df.columns)
                    class_df['Status'] = df['Status'].map(get_status_class)
                    styler = styler.set_td_classes(class_df)

                html_sections[section_name.replace('_', ' ')] = {
                    'df': df,
                    'html': styler.to_html(classes='table table-striped', index=True, border=0),
                    'status': section_status.replace('status-', '')
                }
        
        error_rate = summary_dict.get('Error Rate (%)', 'N/A')

        html_content = template.render(
            tc_id=tc_id,
            description=description,
            overall_status=summary_dict.get('Overall Validation Status', 'N/A'),
            total_records_initial=summary_dict.get('Total Records (Initial)', 'N/A'),
            total_records=summary_dict.get('Total Records (Post-SQL Filter)', 'N/A'),
            records_passed=records_fully_passed,
            records_with_errors=records_with_errors,
            error_rate=error_rate,
            html_sections=html_sections,
            chart_data=chart_data
        )

        html_filepath = os.path.splitext(final_output_file)[0] + '.html'
        with open(html_filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"--- [{tc_id}] HTML report generated at: {html_filepath}")
        return html_filepath
    except Exception as e:
        print(f"{Colors.WARNING}--- [{tc_id}] HTML Report generation failed (Template might be missing): {e}{Colors.ENDC}")
        return None

def apply_scenario_validation(data_df, scenarios_df, tc_id, allowed_val_sep='|'):
    scenario_issues = []
    newly_failed_mask = pd.DataFrame(False, index=data_df.index, columns=data_df.columns)

    if scenarios_df.empty: return scenario_issues, newly_failed_mask

    for _, scenario in scenarios_df.iterrows():
        filter_conditions = str(scenario.get('Filter_Conditions', '')).strip()
        validation_col = str(scenario.get('Validation_Column', '')).strip()
        validation_rule = str(scenario.get('Validation_Rule', '')).strip().upper()
        validation_value = str(scenario.get('Validation_Value', ''))
        
        if not all([filter_conditions, validation_col, validation_rule]): continue
        if validation_col not in data_df.columns: continue

        current_filter_mask = pd.Series(True, index=data_df.index)
        for condition in filter_conditions.split(';'):
            if '=' not in condition: continue
            col, val_str = condition.split('=', 1)
            invert = False
            if val_str.startswith('!'): invert = True; val_str = val_str[1:]
            vals = [v.strip().lower() for v in val_str.split(allowed_val_sep)]
            
            if col not in data_df.columns: current_filter_mask = pd.Series(False, index=data_df.index); break
            condition_mask = data_df[col].astype(str).str.lower().isin(vals)
            if invert: condition_mask = ~condition_mask
            current_filter_mask &= condition_mask
        
        if not current_filter_mask.any(): continue

        data_to_validate = data_df.loc[current_filter_mask, validation_col]
        failure_mask = pd.Series(False, index=data_to_validate.index)

        if validation_rule == 'MANDATORY': failure_mask = data_to_validate.isna() | (data_to_validate.astype(str).str.strip() == '')
        elif validation_rule == 'EMPTY': failure_mask = data_to_validate.notna() & (data_to_validate.astype(str).str.strip() != '')
        elif validation_rule == 'ALLOWED_VALUES':
            allowed_list = [v.strip().lower() for v in validation_value.split(allowed_val_sep)]
            failure_mask = ~data_to_validate.astype(str).str.lower().isin(allowed_list)
        elif validation_rule == 'REGEX': failure_mask = ~data_to_validate.astype(str).str.match(validation_value, na=False)

        if failure_mask.any():
            failed_indices = failure_mask[failure_mask].index
            newly_failed_mask.loc[failed_indices, validation_col] = True
            for index in failed_indices:
                details = f"Failed Scenario '{scenario['Scenario_ID']}': {scenario['Description']}"
                scenario_issues.append({'Row Index': index + 1, 'Column Name': validation_col, 'Value': data_df.at[index, validation_col], 'Validation Rule': 'Scenario', 'Details': details})

    return scenario_issues, newly_failed_mask

def apply_sql_filter_validation(data_df, sql_filter_file, tc_id):
    if pd.isna(sql_filter_file) or not sql_filter_file.strip(): return data_df.copy(), pd.DataFrame()

    sql_path = os.path.join(SQL_RULES_DIR, sql_filter_file)
    if not os.path.exists(sql_path): return data_df.copy(), pd.DataFrame()

    print(f"--- [{tc_id}] Applying SQL filter from: {sql_filter_file}")
    
    filter_params = {}
    try:
        with open(sql_path, 'r') as f: exec(f.read(), {}, filter_params)
    except Exception: return data_df.copy(), pd.DataFrame()
    
    try:
        measure_code = filter_params.get('MEASURE_CODE', filter_params.get('MEASURE_', filter_params.get('Measure')))
        code_exclude = set(filter_params.get('MASTER_EXCLUDE_LIST', filter_params.get('CPT_EXCLUDE', filter_params.get('Code', []))))
        
        data_cols_upper = {col.upper(): col for col in data_df.columns}
        measure_col = data_cols_upper.get('MEASURE')
        code_col = data_cols_upper.get('CODE')
        
        if not measure_code or not code_exclude or not measure_col or not code_col: return data_df.copy(), pd.DataFrame()

        exclusion_mask = (data_df[measure_col].astype(str).str.upper().str.strip() == measure_code.upper()) & \
                         (data_df[code_col].astype(str).str.upper().str.strip().isin(code_exclude))
        
        retained_df = data_df[~exclusion_mask].copy()
        excluded_df = data_df[exclusion_mask].copy()
        print(f"--- [{tc_id}] SQL Filter: {len(retained_df)} rows retained, {len(excluded_df)} rows excluded.")
        return retained_df, excluded_df

    except Exception: return data_df.copy(), pd.DataFrame()

def run_validation_for_test_case(tc_id, description, path, file_type, delimiter, encoding, config_df, scenarios_df, global_settings, sql_filter_file):
    start_time = datetime.now()
    print(f"{Colors.BOLD}--- [{tc_id}] Starting Validation using rules from sheet '{tc_id}' ---{Colors.ENDC}")

    if not os.path.exists(path):
        print(f"{Colors.FAIL}--- [{tc_id}] ERROR: Data File not found at: {path}{Colors.ENDC}")
        print(f"{Colors.WARNING}--- [{tc_id}] Skipping report generation.{Colors.ENDC}")
        return
    else:
        print(f"--- [{tc_id}] Data File: {path}")

    results, failed_row_indices, failed_cell_coordinates = {}, set(), set()
    
    TOP_N_DATA = global_settings.get('TOP_N_DATA', 5)
    
    enable_cell_highlighting = str(global_settings.get('Enable_Cell_Highlighting', 'Yes')).lower() in ['yes', 'true', '1']
    enable_column_autofit = str(global_settings.get('Enable_Column_Autofit', 'Yes')).lower() in ['yes', 'true', '1']
    auto_open = str(global_settings.get('Auto_Open_Report', 'Yes')).lower() in ['yes', 'true', '1']
    generate_passed = str(global_settings.get('Generate_Passed_Data_Report', 'True')).lower() in ['yes', 'true', '1']
    allowed_val_sep = str(global_settings.get('Allowed_Values_Separator', '|'))

    duplicate_key_columns = [c.strip() for c in str(global_settings.get('Duplicate_Key_Columns', '')).split(',') if c.strip()]

    config_df['Column_Name'] = config_df['Column_Name'].astype(str).str.strip()
    config_df['Data_Type'] = config_df['Data_Type'].astype(str).str.strip().str.lower()
    config_df['Use'] = config_df['Use'].astype(str).str.strip().str.lower()
    config_df['Length'] = config_df['Length'].astype(str)
    config_df['Dateformat'] = config_df['Dateformat'].apply(clean_dateformat)
    expected_data_columns = config_df['Column_Name'].tolist()
    VALID_DATA_TYPES = {'int', 'bigint', 'numeric', 'float', 'decimal', 'string', 'varchar', 'nchar', 'date', 'boolean', 'bit'}

    config_issues = []
    for index, row in config_df.iterrows():
        if row['Data_Type'] not in VALID_DATA_TYPES:
            config_issues.append({'Column Name': row['Column_Name'], 'Problem': f"Invalid Data_Type '{row['Data_Type']}'", 'Details': 'Invalid Type', 'Status': 'FAIL'})
    results['Configuration_Issues'] = pd.DataFrame(config_issues)

    structure_results, data_df, found_data_columns = [], pd.DataFrame(), []
    
    if os.path.exists(path):
        structure_results.append({'Test': 'File Exists', 'Expected': 'Yes', 'Actual': 'Yes', 'Status': 'PASS'})
        try:
            if file_type in ['csv', 'txt']:
                read_csv_kwargs = {'delimiter': delimiter, 'dtype': str, 'keep_default_na': False, 'encoding': encoding, 'engine': 'c', 'on_bad_lines': 'skip', 'quotechar': '"', 'skipinitialspace': True}
                try: data_df = pd.read_csv(path, **read_csv_kwargs)
                except Exception: 
                    read_csv_kwargs['engine'] = 'python'; del read_csv_kwargs['on_bad_lines']
                    data_df = pd.read_csv(path, **read_csv_kwargs)
                found_data_columns = data_df.columns.tolist()
            elif file_type == 'xlsx':
                with pd.ExcelFile(path) as xls_data:
                    data_df = pd.read_excel(xls_data, header=0, dtype=str, keep_default_na=False)
                found_data_columns = data_df.columns.tolist()
            else: raise ValueError(f"Unsupported file type: {file_type}")

            data_df.columns = [str(col).strip() for col in found_data_columns]
            found_data_columns = data_df.columns.tolist()
            for col in data_df.columns:
                if data_df[col].dtype == 'object': data_df[col] = data_df[col].fillna('').astype(str).str.strip().str.strip('"')
            
            structure_results.append({'Test': 'File Read Successfully', 'Expected': 'Yes', 'Actual': 'Yes', 'Status': 'PASS'})
            structure_results.append({'Test': 'Record Count', 'Expected': '>= 0', 'Actual': len(data_df), 'Status': 'INFO'})
        except Exception as e:
            structure_results.append({'Test': 'File Read Successfully', 'Expected': 'Yes', 'Actual': f'No - Error: {e}', 'Status': 'FAIL'})
    else:
        structure_results.append({'Test': 'File Exists', 'Expected': 'Yes', 'Actual': 'No', 'Status': 'FAIL'})

    results['Structure'] = pd.DataFrame(structure_results)
    results['TestData_Initial'] = data_df.copy()
    initial_record_count = len(data_df)
    
    sql_excluded_df = pd.DataFrame()
    if not data_df.empty:
        data_df, sql_excluded_df = apply_sql_filter_validation(data_df, sql_filter_file, tc_id)
        if not sql_excluded_df.empty: sql_excluded_df.insert(0, 'SQL File', sql_filter_file)
    results['SQL_Exclusions'] = sql_excluded_df
    results['TestData_PostFilter'] = data_df.copy()

    if not data_df.empty:
        temp_length_col = f"__temp_row_len_{datetime.now().timestamp()}__"
        data_df[temp_length_col] = data_df.apply(lambda x: len(''.join(x.astype(str))), axis=1)
        results['Maximum_Data'] = data_df.nlargest(TOP_N_DATA, temp_length_col).drop(columns=[temp_length_col])
        results['Minimum_Data'] = data_df.nsmallest(TOP_N_DATA, temp_length_col).drop(columns=[temp_length_col])
        data_df = data_df.drop(columns=[temp_length_col])
    else: results['Maximum_Data'], results['Minimum_Data'] = pd.DataFrame(), pd.DataFrame()

    config_set, found_set = set(expected_data_columns), set(found_data_columns)
    results['Column_Comparison'] = pd.DataFrame([
        {'Comparison': 'Missing Expected Columns', 'Details': ', '.join(config_set - found_set) or 'N/A', 'Status': 'FAIL' if config_set - found_set else 'PASS'},
        {'Comparison': 'Unexpected Columns Found', 'Details': ', '.join(found_set - config_set) or 'N/A', 'Status': 'WARN' if found_set - config_set else 'PASS'}
    ])
    
    col_seq_df = pd.DataFrame([{'Sno': i + 1, 'Configuration': expected_data_columns[i] if i < len(expected_data_columns) else 'N/A', 'File': found_data_columns[i] if i < len(found_data_columns) else 'N/A', 'Status': ('PASS' if (i < len(expected_data_columns) and i < len(found_data_columns) and expected_data_columns[i] == found_data_columns[i]) else 'FAIL')} for i in range(max(len(expected_data_columns), len(found_data_columns)))])
    results['Column_Sequence'] = col_seq_df

    whitespace_issues, length_issues, datatype_issues, datacontent_issues, dateformat_issues, content_issues = [], [], [], [], [], []
    if not data_df.empty:
        for i, col_name in enumerate(found_data_columns):
            if col_name != col_name.strip(): whitespace_issues.append({'Row Index': 'Header', 'Column Name': f'Column {i+1} ({col_name})', 'Value': f'"{col_name}"', 'Details': 'Header contains whitespace'})
        for col_name in data_df.columns:
            if data_df[col_name].dtype == 'object':
                mask = (data_df[col_name].notna()) & (data_df[col_name].str.strip() != data_df[col_name])
                for index in data_df[col_name][mask].index: whitespace_issues.append({'Row Index': index + 1, 'Column Name': col_name, 'Value': f'"{data_df.at[index, col_name]}"', 'Details': 'Cell has whitespace'})
    results['WhiteSpace'] = pd.DataFrame(whitespace_issues)

    if not data_df.empty:
        failed_mask_df = pd.DataFrame(False, index=data_df.index, columns=data_df.columns)
        
        for _, config_row in config_df.iterrows():
            col_name = config_row['Column_Name']
            if col_name not in data_df.columns: continue
            column_series = data_df[col_name]
            
            if config_row['Use'] == 'mandatory':
                mask = column_series.isna() | (column_series.astype(str).str.strip() == "")
                if mask.any():
                    failed_mask_df.loc[mask, col_name] = True
                    for idx in mask[mask].index: datacontent_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': '', 'Details': 'Mandatory Empty'})
            
            not_empty_mask = ~(column_series.isna() | (column_series.astype(str).str.strip() == ""))
            working_series = column_series[not_empty_mask]
            if working_series.empty: continue

            length_val = str(config_row.get('Length', 'nan'))
            if length_val != 'nan' and config_row['Data_Type'] != 'date':
                str_lens = working_series.astype(str).str.len()
                mask = pd.Series(False, index=working_series.index)
                rule = str(config_row.get('Length_Rule', 'MAX')).upper()
                try:
                    if rule == 'EQUAL': mask = str_lens != int(length_val)
                    elif rule == 'RANGE': mn, mx = map(int, length_val.split('-')); mask = (str_lens < mn) | (str_lens > mx)
                    else: mask = str_lens > int(length_val)
                    if mask.any():
                        failed_mask_df.loc[mask.index[mask], col_name] = True
                        for idx in mask[mask].index: length_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': data_df.at[idx, col_name], 'Actual Length': str_lens[idx], 'Expected Rule': rule})
                except: pass

            d_type = config_row['Data_Type']
            if d_type in ['int', 'bigint']:
                mask = ~working_series.astype(str).str.match(r'^-?\d+$')
                if mask.any():
                    failed_mask_df.loc[mask.index[mask], col_name] = True
                    for idx in mask[mask].index: datatype_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': data_df.at[idx, col_name], 'Expected Type': d_type})
            
            elif d_type == 'date':
                fmt = str(config_row.get('Dateformat', 'nan')).strip()
                date_range_str = str(config_row.get('Date_Range', 'nan')).strip()
                allow_future = str(config_row.get('Allow_Future_Date', 'No')).strip().lower() == 'yes'

                parsed = pd.to_datetime(working_series, format=fmt, errors='coerce')
                mask = parsed.isna()
                
                if mask.any():
                    failed_mask_df.loc[mask.index[mask], col_name] = True
                    for idx in mask[mask].index: dateformat_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': data_df.at[idx, col_name], 'Expected Format': fmt, 'Details': 'Format Mismatch'})

                valid_parsed = parsed.dropna()
                if not valid_parsed.empty:
                    if not allow_future:
                        future_mask = valid_parsed.dt.date > datetime.now().date()
                        if future_mask.any():
                            failed_mask_df.loc[future_mask.index[future_mask], col_name] = True
                            for idx in future_mask[future_mask].index: dateformat_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': data_df.at[idx, col_name], 'Expected Format': fmt, 'Details': 'Future Date Not Allowed'})

                    if date_range_str != 'nan':
                        try:
                            if '/' in date_range_str:
                                min_s, max_s = date_range_str.split('/')
                                min_d = datetime.strptime(min_s, '%Y-%m-%d').date()
                                max_d = datetime.strptime(max_s, '%Y-%m-%d').date()
                                range_mask = (valid_parsed.dt.date < min_d) | (valid_parsed.dt.date > max_d)
                            else:
                                min_d = datetime.strptime(date_range_str, '%Y-%m-%d').date()
                                range_mask = valid_parsed.dt.date < min_d
                            
                            if range_mask.any():
                                failed_mask_df.loc[range_mask.index[range_mask], col_name] = True
                                for idx in range_mask[range_mask].index: dateformat_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': data_df.at[idx, col_name], 'Expected Format': fmt, 'Details': f'Date out of range ({date_range_str})'})
                        except ValueError: pass 

            if 'Regex_Pattern' in config_row and pd.notna(config_row['Regex_Pattern']):
                try:
                    mask = ~working_series.astype(str).str.match(str(config_row['Regex_Pattern']).strip(), na=False)
                    if mask.any():
                        failed_mask_df.loc[mask.index[mask], col_name] = True
                        for idx in mask[mask].index: content_issues.append({'Row Index': idx + 1, 'Column Name': col_name, 'Value': data_df.at[idx, col_name], 'Validation Rule': 'Regex', 'Details': 'Pattern Mismatch'})
                except: pass

        scenario_errors, scenario_failed_mask = apply_scenario_validation(data_df, scenarios_df, tc_id, allowed_val_sep)
        content_issues.extend(scenario_errors)
        failed_mask_df |= scenario_failed_mask
        
        failed_row_indices = {idx for idx, row_failed in failed_mask_df.any(axis=1).items() if row_failed}
        for (row_idx, col_name), is_failed in failed_mask_df.stack().items():
            if is_failed: failed_cell_coordinates.add((row_idx, col_name, data_df.columns.get_loc(col_name)))
    
    duplicate_df = pd.DataFrame()
    if not data_df.empty:
        if duplicate_key_columns and all(col in data_df.columns for col in duplicate_key_columns):
            duplicate_df = data_df[data_df.duplicated(subset=duplicate_key_columns, keep=False)].copy()
            if not duplicate_df.empty: duplicate_df.insert(0, 'Duplicate Key Columns', ', '.join(duplicate_key_columns))
        else:
            duplicate_df = data_df[data_df.duplicated(keep=False)].copy()
            if not duplicate_df.empty: duplicate_df.insert(0, 'Duplicate Key Columns', 'All Columns')

    passed_data_df = pd.DataFrame()
    if generate_passed and not data_df.empty:
        passed_data_df = data_df.drop(index=list(failed_row_indices)).copy()

    results.update({'Data_Content': pd.DataFrame(datacontent_issues), 'Length_Compare': pd.DataFrame(length_issues), 'DataType': pd.DataFrame(datatype_issues), 'Date_Format': pd.DataFrame(dateformat_issues), 'Content_Validation': pd.DataFrame(content_issues), 'Data_Failed': data_df.loc[sorted(list(failed_row_indices))].copy(), 'Passed_Data': passed_data_df, 'Duplicate': duplicate_df, 'Non_Functional': pd.DataFrame([{'Test': 'Performance', 'Status': 'NOT TESTED'}])})
    
    results['Error_Summary'] = generate_error_summary(results)
    is_fail = bool(failed_row_indices or any(s['Status'] == 'FAIL' for s in structure_results) or any(c['Status'] == 'FAIL' for c in results['Column_Comparison'].to_dict('records')) or not results['WhiteSpace'].empty or not results['Configuration_Issues'].empty or not results['SQL_Exclusions'].empty)
    overall_status = 'FAIL' if is_fail else 'PASS'

    total_records_processed = len(data_df)
    records_with_errors = len(failed_row_indices)
    error_rate = (records_with_errors / total_records_processed * 100) if total_records_processed > 0 else 0
    
    results['Summary'] = pd.DataFrame({'Metric': ['--- Overall Result ---', 'Overall Validation Status', '','--- File Details ---', 'Source File', 'Rule Sheet Name', 'Total Records (Initial)', 'Total Records (Post-SQL Filter)', '','--- Validation Summary ---', 'Records Passed', 'Records with Errors', 'Error Rate (%)'], 'Value': ['', overall_status, '','', os.path.basename(path), tc_id, initial_record_count, total_records_processed, '','', total_records_processed - records_with_errors, records_with_errors, f"{error_rate:.2f}%"]})

    output_folder = global_settings.get('Output_folder')
    base_name = re.sub(r'[<>:"/\\|?*]', '_', description).strip().replace(' ', '_') or tc_id
    final_output_file = get_unique_filepath(os.path.join(output_folder, f"{base_name}.xlsx"))

    print(f"--- [{tc_id}] Writing results to: {final_output_file}")

    try:
        with pd.ExcelWriter(final_output_file, engine='openpyxl') as writer:
            sheet_config = {
                'Summary': {'df': results['Summary'], 'index': False}, 'Error_Summary': {'df': results['Error_Summary'], 'index': False},
                'Configuration_Issues': {'df': results['Configuration_Issues'], 'index': True}, 'Structure': {'df': results['Structure'], 'index': True},
                'Column_Comparison': {'df': results['Column_Comparison'], 'index': True}, 'Column_Sequence': {'df': results['Column_Sequence'], 'index': False},
                'WhiteSpace': {'df': results['WhiteSpace'], 'index': False}, 'Data_Content': {'df': results['Data_Content'], 'index': False},
                'DataType': {'df': results['DataType'], 'index': False}, 'Date_Format': {'df': results['Date_Format'], 'index': False},
                'Length_Compare': {'df': results['Length_Compare'], 'index': False}, 'Content_Validation': {'df': results['Content_Validation'], 'index': False},
                'SQL_Exclusions': {'df': results['SQL_Exclusions'], 'index': True, 'headers': found_data_columns},
                'Data_Failed': {'df': results['Data_Failed'], 'index': True, 'headers': found_data_columns}, 
                'Passed_Data': {'df': results['Passed_Data'], 'index': True, 'headers': found_data_columns},
                'Duplicate': {'df': results['Duplicate'], 'index': True, 'headers': found_data_columns},
                'Minimum_Data': {'df': results['Minimum_Data'], 'index': True}, 'Maximum_Data': {'df': results['Maximum_Data'], 'index': True},
                'TestData_Initial': {'df': results['TestData_Initial'], 'index': True}, 'TestData_PostFilter': {'df': results['TestData_PostFilter'], 'index': True}
            }
            
            for sheet_name, config in sheet_config.items():
                df = config.get('df')
                if df is None or (df.empty and sheet_name not in ['Summary', 'Structure', 'Column_Comparison']): continue
                
                final_headers = config.get('headers', df.columns.tolist())
                write_index = config.get('index', False)
                df.to_excel(writer, sheet_name=sheet_name, index=write_index, header=True)
                format_and_autofit_sheet(writer.sheets[sheet_name], df, final_headers, write_index, enable_autofit=enable_column_autofit)

            if enable_cell_highlighting:
                highlight_overall_status(writer, overall_status)
                if 'Data_Failed' in writer.sheets: highlight_failed_cells(writer, results['Data_Failed'], failed_cell_coordinates)
                for s in ['Data_Content', 'DataType', 'Date_Format', 'Length_Compare', 'Content_Validation', 'WhiteSpace', 'Duplicate']:
                    if s in writer.sheets: highlight_detail_rows(writer, s)

        html_report_path = generate_html_report(tc_id, description, results, final_output_file)
        
        if auto_open and html_report_path:
            try:
                if sys.platform == "win32": os.startfile(html_report_path)
                elif sys.platform == "darwin": subprocess.run(["open", html_report_path])
                else: subprocess.run(["xdg-open", html_report_path])
            except Exception: pass

    except Exception as e:
        print(f"{Colors.FAIL}--- [{tc_id}] ERROR: Report writing failed: {e}{Colors.ENDC}"); traceback.print_exc()

def main_validation_logic(config_file):
    print(f"{Colors.HEADER}{Colors.BOLD}--- Starting Data Validation (Manual Mode) ---{Colors.ENDC}")
    print(f"Configuration File: {config_file}\n")

    if not os.path.exists(config_file):
        print(f"{Colors.FAIL}ERROR: Configuration.xlsx not found at: {config_file}{Colors.ENDC}")
        return

    try:
        xls = pd.ExcelFile(config_file)
        all_sheet_names = xls.sheet_names
        global_settings_df = xls.parse('Settings')
        global_settings = dict(zip(global_settings_df['Setting'], global_settings_df['Value']))
        
        run_df = xls.parse('Run')
        run_df.columns = run_df.columns.str.strip()
    except Exception as e:
        print(f"{Colors.FAIL}ERROR: Could not load configuration sheets. Details: {e}{Colors.ENDC}")
        return

    output_folder = global_settings.get('Output_folder')
    if not output_folder or pd.isna(output_folder):
        print(f"{Colors.FAIL}CRITICAL ERROR: 'Output_folder' not defined in Settings.{Colors.ENDC}")
        return

    if not os.path.exists(output_folder): os.makedirs(output_folder)
    if not os.path.exists(SQL_RULES_DIR): os.makedirs(SQL_RULES_DIR)

    run_df = run_df[run_df['Execute'].astype(str).str.strip().str.lower() == 'yes'].copy()
    
    if run_df.empty:
        print(f"{Colors.WARNING}--- WARNING: No Test Cases marked 'Yes' for Execute in the Run sheet. ---{Colors.ENDC}")
        return
    
    print(f"{Colors.OKGREEN}Found {len(run_df)} Test Case(s) to execute.{Colors.ENDC}")

    for i, (_, test_run) in enumerate(run_df.iterrows()):
        tc_id = test_run['Test Case']
        description = test_run['Description']
        
        print("-" * 50)
        print(f"{Colors.OKCYAN}Processing {i+1}/{len(run_df)}: {tc_id} - {description}{Colors.ENDC}")

        file_path = str(test_run['Path'])

        rule_sheet_name = tc_id
        if rule_sheet_name not in all_sheet_names:
            print(f"{Colors.FAIL}ERROR: Rule sheet '{rule_sheet_name}' not found. Skipping.{Colors.ENDC}")
            continue
        
        config_df = pd.read_excel(xls, sheet_name=rule_sheet_name)
        scenarios_df = pd.DataFrame()
        if f"{tc_id}_Scenarios" in all_sheet_names: scenarios_df = pd.read_excel(xls, sheet_name=f"{tc_id}_Scenarios")

        encoding_val = test_run.get('Encoding', 'utf-8-sig')
        file_encoding = 'utf-8-sig' if pd.isna(encoding_val) or str(encoding_val).strip() == '' else str(encoding_val).strip()
        sql_filter_file = test_run.get('SQL_Filter_File')

        run_validation_for_test_case(
            tc_id=tc_id, description=description, path=file_path,
            file_type=str(test_run['File_Type']).lower(), delimiter=str(test_run['Delimiter']),
            encoding=file_encoding, config_df=config_df, scenarios_df=scenarios_df,
            global_settings=global_settings, sql_filter_file=sql_filter_file
        )

    print("\n" + "=" * 50)
    print(f"{Colors.HEADER}--- All Validation Tasks Completed ---{Colors.ENDC}")

if __name__ == "__main__":
    clear_screen()
    if os.path.exists(CONFIG_FILE):
        main_validation_logic(CONFIG_FILE)
    else:
        print(f"{Colors.FAIL}CRITICAL: Configuration.xlsx not found in {SCRIPT_DIR}{Colors.ENDC}")
